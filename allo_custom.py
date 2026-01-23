import numpy as np
import pandas as pd
import streamlit as st
from scipy.optimize import minimize

st.set_page_config(
    page_title="[글로벌자산배분전략위원회] Allocation",
    layout="wide",
)

# 사이드바 텍스트 크기 조정 CSS
st.markdown("""
    <style>
        /* 사이드바 전체 텍스트 크기 조정 */
        [data-testid="stSidebar"] {
            font-size: 14px !important;
        }

        /* 사이드바 제목 */
        [data-testid="stSidebar"] h1,
        [data-testid="stSidebar"] h2,
        [data-testid="stSidebar"] h3 {
            font-size: 16px !important;
        }

        /* 사이드바 본문 텍스트 */
        [data-testid="stSidebar"] p,
        [data-testid="stSidebar"] div,
        [data-testid="stSidebar"] span {
            font-size: 14px !important;
        }

        /* 사이드바 입력 필드 */
        [data-testid="stSidebar"] input,
        [data-testid="stSidebar"] select,
        [data-testid="stSidebar"] textarea {
            font-size: 14px !important;
        }

        /* 사이드바 라디오 버튼, 체크박스 */
        [data-testid="stSidebar"] label {
            font-size: 14px !important;
        }

        /* 사이드바 버튼 */
        [data-testid="stSidebar"] button {
            font-size: 14px !important;
        }

        /* 사이드바 메트릭 */
        [data-testid="stSidebar"] [data-testid="stMetricValue"] {
            font-size: 15px !important;
        }

        [data-testid="stSidebar"] [data-testid="stMetricLabel"] {
            font-size: 13px !important;
        }
    </style>
""", unsafe_allow_html=True)


def normalize_percentage(value):
    """Convert Excel-style percentage to decimal."""
    if pd.isna(value):
        return np.nan

    if isinstance(value, str):
        cleaned = value.replace("%", "").replace(",", "").strip()
        if cleaned == "":
            return np.nan
        try:
            value = float(cleaned)
        except ValueError:
            return np.nan

    try:
        value = float(value)
    except (TypeError, ValueError):
        return np.nan

    return value / 100 if value > 1.5 else value


def load_excel(uploaded_file):
    """Read Excel file."""
    price_df = pd.read_excel(uploaded_file, sheet_name="기준가")
    current_df = pd.read_excel(uploaded_file, sheet_name="Current")
    asset_minmax_df = pd.read_excel(uploaded_file, sheet_name="Asset_MinMax")
    group_df = pd.read_excel(uploaded_file, sheet_name="Gr_MinMax")

    # BM 시트는 선택사항
    bm_df = None
    try:
        bm_df = pd.read_excel(uploaded_file, sheet_name="BM")
    except:
        pass

    return price_df, current_df, asset_minmax_df, group_df, bm_df


def prepare_current_df(df):
    """Clean Current sheet."""
    clean_df = df.copy()
    for col in ["CURRENT", "MIN", "MAX", "EXPECTED_R", "YLD"]:
        if col in clean_df.columns:
            if col == "EXPECTED_R":
                # EXPECTED_R은 특별 처리: 5.0 또는 500을 "기대값 없음"으로 인식
                # 원본 값을 먼저 확인
                original_values = clean_df[col].copy()
                clean_df[col] = clean_df[col].apply(normalize_percentage)
                # 5.0 또는 500을 "기대값 없음"으로 인식하여 그대로 유지
                mask_5 = np.isclose(original_values, 5.0, atol=0.1) | np.isclose(original_values, 500.0, atol=1.0)
                # normalize_percentage 적용 후 0.05가 된 경우를 5.0으로 복원
                mask_normalized = np.isclose(clean_df[col], 0.05, atol=0.01) & mask_5
                clean_df.loc[mask_normalized, col] = 5.0
            else:
                clean_df[col] = clean_df[col].apply(normalize_percentage)
    return clean_df


def prepare_group_df(df, gd_option="GD"):
    """Clean group constraints."""
    group_df = df.copy()
    # GD 옵션에 따라 처리할 컬럼 결정
    if gd_option == "GD":
        cols_to_normalize = ["Min_GD", "Max_GD"]
    else:  # GD+
        cols_to_normalize = ["Min_GDp", "Max_GDp"]
    
    for col in cols_to_normalize:
        if col in group_df.columns:
            group_df[col] = group_df[col].apply(normalize_percentage)
    return group_df


def merge_asset_data(current_df, asset_minmax_df, group_df, gd_option):
    """Merge Current, Asset_MinMax, and Gr_MinMax data based on GD option."""
    # GD 옵션에 따라 컬럼명 결정
    if gd_option == "GD":
        asset_min_col = "Min_GD"
        asset_max_col = "Max_GD"
        group_min_col = "Min_GD"
        group_max_col = "Max_GD"
    else:  # GD+
        asset_min_col = "Min_GDp"
        asset_max_col = "Max_GDp"
        group_min_col = "Min_GDp"
        group_max_col = "Max_GDp"
    
    # Current와 Asset_MinMax 병합 (CODE, NAME을 key로)
    # Current 시트에 이미 GROUP이 있으므로 Min_GD/Max_GD 또는 Min_GDp/Max_GDp만 가져옴
    merged_df = current_df.merge(
        asset_minmax_df[["CODE", "NAME", asset_min_col, asset_max_col]],
        on=["CODE", "NAME"],
        how="left"
    )
    
    # Min_GD/Max_GD 또는 Min_GDp/Max_GDp를 Min, Max로 이름 변경
    merged_df = merged_df.rename(columns={
        asset_min_col: "MIN",
        asset_max_col: "MAX"
    })
    
    # GROUP으로 정렬
    merged_df = merged_df.sort_values("GROUP").reset_index(drop=True)
    
    # GROUP을 key로 Gr_MinMax 병합
    group_merge_cols = ["GROUP", group_min_col, group_max_col]
    if all(col in group_df.columns for col in group_merge_cols):
        merged_df = merged_df.merge(
            group_df[group_merge_cols],
            on="GROUP",
            how="left"
        )
        # Gr_MinMax 컬럼 이름 변경
        merged_df = merged_df.rename(columns={
            group_min_col: "Gr_Min",
            group_max_col: "Gr_Max"
        })
    else:
        # 컬럼이 없으면 빈 값으로 추가
        merged_df["Gr_Min"] = np.nan
        merged_df["Gr_Max"] = np.nan
    
    return merged_df


def format_percentage_int(value):
    """Format as integer percentage (no decimals)."""
    if pd.isna(value):
        return ""
    return f"{int(value * 100)}%"


def format_percentage_1dec(value):
    """Format as percentage with 1 decimal."""
    if pd.isna(value):
        return ""
    return f"{value * 100:.1f}%"


def format_percentage_2dec(value):
    """Format as percentage with 2 decimals, or 'X' if value is 5.0."""
    if pd.isna(value):
        return ""
    # 5.0 (500%)인 경우 X로 표시
    if np.isclose(value, 5.0, atol=0.1):
        return "X"
    return f"{value * 100:.2f}%"


def format_display_df(df, raw_expected_r=None):
    """Format dataframe for display with special formatting."""
    display_df = df.copy()
    
    # CODE 컬럼 제거
    if "CODE" in display_df.columns:
        display_df = display_df.drop(columns=["CODE"])
    
    # CURRENT를 소숫점 첫째자리 퍼센트로 변환
    if "CURRENT" in display_df.columns:
        display_df["CURRENT"] = display_df["CURRENT"].apply(format_percentage_1dec)
    
    # Min, Max, Gr_Min, Gr_Max를 정수 퍼센트로 변환
    for col in ["MIN", "MAX", "Gr_Min", "Gr_Max"]:
        if col in display_df.columns:
            display_df[col] = display_df[col].apply(format_percentage_int)
    
    # YLD를 소숫점 둘째자리 퍼센트로 변환
    if "YLD" in display_df.columns:
        display_df["YLD"] = display_df["YLD"].apply(format_percentage_2dec)
    
    # EXPECTED_R을 소숫점 둘째자리 퍼센트로 변환 (5.0이면 X)
    if "EXPECTED_R" in display_df.columns:
        display_df["EXPECTED_R"] = display_df["EXPECTED_R"].apply(format_percentage_2dec)
    
    # Gr_Min, Gr_Max는 각 GROUP의 첫 행에만 표시
    if "Gr_Min" in display_df.columns and "GROUP" in display_df.columns:
        if len(display_df) > 0:
            # 첫 행은 항상 첫 그룹의 첫 행
            display_df["_is_first_in_group"] = False
            display_df.loc[0, "_is_first_in_group"] = True
            # 나머지 행들은 GROUP이 바뀌는 지점
            if len(display_df) > 1:
                display_df.loc[1:, "_is_first_in_group"] = display_df.loc[1:, "GROUP"].values != display_df.loc[:len(display_df)-2, "GROUP"].values
            display_df.loc[~display_df["_is_first_in_group"], "Gr_Min"] = ""
            display_df.loc[~display_df["_is_first_in_group"], "Gr_Max"] = ""
            display_df = display_df.drop(columns=["_is_first_in_group"])
    
    # 칼럼명 변경
    column_rename = {
        "NAME": "Sleeve",
        "CURRENT": "현재비중",
        "EXPECTED_R": "기대수익률"
    }
    display_df = display_df.rename(columns=column_rename)
    
    return display_df


def validate_codes(price_df, codes):
    """Ensure each code exists in price sheet."""
    missing = [code for code in codes if code not in price_df.columns]
    if missing:
        raise ValueError(
            f"기준가 시트에 존재하지 않는 CODE: {', '.join(missing)}"
        )


def compute_monthly_returns(price_df, codes, months):
    """Build monthly return matrix."""
    data = price_df.copy()
    if "DATE" not in data.columns:
        raise ValueError("기준가 시트에 DATE 컬럼이 없습니다.")

    data["DATE"] = pd.to_datetime(data["DATE"], errors="coerce")
    data = data.sort_values("DATE")
    data = data.set_index("DATE")

    missing_cols = [code for code in codes if code not in data.columns]
    if missing_cols:
        raise ValueError(f"CODE에 해당하는 기준가 컬럼이 없습니다: {missing_cols}")

    price_matrix = data[codes].ffill().dropna(how="all")
    monthly_price = price_matrix.resample("M").last()
    monthly_returns = monthly_price.pct_change().dropna(how="all")
    monthly_returns = monthly_returns.tail(months)

    if monthly_returns.empty:
        raise ValueError("선택한 참조기간에 활용 가능한 월별 수익률이 없습니다.")

    return monthly_returns


def compute_bm_returns(bm_df, months):
    """Build BM monthly return series from BM sheet."""
    if bm_df is None:
        return None

    if "DATE" not in bm_df.columns or "TR" not in bm_df.columns:
        return None

    data = bm_df.copy()
    data["DATE"] = pd.to_datetime(data["DATE"], errors="coerce")
    data = data.sort_values("DATE")
    data = data.set_index("DATE")

    # TR 값이 모두 0인지 확인
    if data["TR"].fillna(0).abs().sum() < 1e-10:
        return None

    bm_price = data[["TR"]].ffill().dropna(how="all")
    monthly_bm_price = bm_price.resample("M").last()
    monthly_bm_returns = monthly_bm_price.pct_change().dropna(how="all")
    monthly_bm_returns = monthly_bm_returns.tail(months)

    if monthly_bm_returns.empty:
        return None

    return monthly_bm_returns["TR"]


def annualize_series(monthly_series):
    """Convert monthly mean returns to annual."""
    return (1 + monthly_series).pow(12) - 1


def annualize_vol(vol_monthly):
    """Convert monthly vol to annual."""
    return vol_monthly * np.sqrt(12)


def build_expected_returns(method, current_df, hist_returns, monthly_means, cov_monthly, sims, raw_expected_r=None,
                           replacement_method=None):
    codes = current_df["CODE"].tolist()
    excel_values = current_df.set_index("CODE")["EXPECTED_R"]
    hist = hist_returns.reindex(excel_values.index)

    if method == "excel":
        result = excel_values.copy()
        # 5.0 또는 500%는 기대값이 없다는 의미
        # 정규화된 값에서 5.0을 체크하거나, 원본 값에서 5.0 또는 500을 체크
        if raw_expected_r is not None:
            raw_aligned = raw_expected_r.reindex(excel_values.index)
            placeholders = result.isna() | np.isclose(result, 5.0, atol=0.1) | (raw_aligned.notna() & (
                        np.isclose(raw_aligned, 5.0, atol=0.1) | np.isclose(raw_aligned, 500.0, atol=1.0)))
        else:
            placeholders = result.isna() | np.isclose(result, 5.0, atol=0.1)
        if placeholders.any():
            # 사용자가 선택한 대체 방식으로 값 대체
            if replacement_method == "monte_carlo":
                # 몬테칼로 시뮬레이션으로 대체
                np.random.seed(42)
                cov = cov_monthly.values + np.eye(len(cov_monthly)) * 1e-8
                try:
                    sims_res = np.random.multivariate_normal(monthly_means.values, cov, size=sims)
                    mc_means = sims_res.mean(axis=0)
                    ann_mc = annualize_series(pd.Series(mc_means, index=monthly_means.index))
                    mc_values = ann_mc.reindex(excel_values.index)
                    result[placeholders] = mc_values[placeholders]
                except np.linalg.LinAlgError:
                    # 공분산 행렬 오류 시 과거수익률로 대체
                    result[placeholders] = hist[placeholders]
            else:
                # 과거수익률로 대체 (default)
                result[placeholders] = hist[placeholders]
    elif method == "historical":
        result = hist.copy()
    else:  # monte_carlo
        np.random.seed(42)
        cov = cov_monthly.values + np.eye(len(cov_monthly)) * 1e-8
        try:
            sims_res = np.random.multivariate_normal(monthly_means.values, cov, size=sims)
            mc_means = sims_res.mean(axis=0)
            ann_mc = annualize_series(pd.Series(mc_means, index=monthly_means.index))
            result = ann_mc.reindex(excel_values.index)
        except np.linalg.LinAlgError:
            result = hist.copy()

    return result.reindex(codes).fillna(0.0)


def build_expected_vols(method, current_df, hist_vols, monthly_means, cov_monthly, sims, raw_expected_v=None,
                        replacement_method=None):
    """Build expected volatilities. Always uses historical volatility since EXPECTED_V is not used."""
    # EXPECTED_V는 사용하지 않으므로 항상 과거 변동성 사용
    return hist_vols.copy()


def risk_parity_objective(w, cov):
    portfolio_var = np.dot(w, cov @ w)
    if portfolio_var <= 0:
        return 1e6
    marginal_contrib = cov @ w
    risk_contrib = w * marginal_contrib
    target = portfolio_var / len(w)
    return np.sum((risk_contrib - target) ** 2)


def sortino_ratio(w, returns_matrix, mar, rf):
    port_monthly = returns_matrix @ w
    mar_monthly = (1 + mar) ** (1 / 12) - 1
    downside = np.minimum(port_monthly - mar_monthly, 0)
    downside_dev = np.sqrt(np.mean(downside ** 2)) if np.any(downside) else 1e-6
    ann_return = (1 + port_monthly.mean()) ** 12 - 1
    ann_downside = downside_dev * np.sqrt(12)
    return (ann_return - rf) / ann_downside


def portfolio_stats(w, mu, cov, dur, yld, returns_matrix, mar, rf):
    port_return = np.dot(w, mu)
    port_vol = np.sqrt(np.dot(w, cov @ w))
    sharpe = (port_return - rf) / port_vol if port_vol > 0 else np.nan
    sortino = sortino_ratio(w, returns_matrix, mar, rf)
    duration = np.dot(w, dur)
    portfolio_yield = np.dot(w, yld)
    return {
        "return": port_return,
        "vol": port_vol,
        "sharpe": sharpe,
        "sortino": sortino,
        "duration": duration,
        "yield": portfolio_yield,
    }


def portfolio_vs_bm_stats(w, mu, cov, returns_matrix, bm_returns_monthly):
    """Calculate portfolio vs benchmark comparison statistics."""
    if bm_returns_monthly is None or len(bm_returns_monthly) == 0:
        return None

    # 포트폴리오 월별 수익률
    port_monthly = returns_matrix @ w

    # 공통 기간 찾기
    common_dates = returns_matrix.index.intersection(bm_returns_monthly.index)
    if len(common_dates) == 0:
        return None

    port_common = port_monthly.loc[common_dates]
    bm_common = bm_returns_monthly.loc[common_dates]

    # 연율화
    port_ann_return = (1 + port_common.mean()) ** 12 - 1
    bm_ann_return = (1 + bm_common.mean()) ** 12 - 1

    # Alpha (초과수익률)
    alpha = port_ann_return - bm_ann_return

    # Tracking Error (연율화)
    active_returns = port_common - bm_common
    tracking_error = np.std(active_returns) * np.sqrt(12)

    # Information Ratio
    ir = alpha / tracking_error if tracking_error > 0 else np.nan

    # 상관관계
    correlation = np.corrcoef(port_common, bm_common)[0, 1] if len(port_common) > 1 else np.nan

    # 베타
    if len(port_common) > 1 and np.var(bm_common) > 0:
        beta = np.cov(port_common, bm_common)[0, 1] / np.var(bm_common)
    else:
        beta = np.nan

    return {
        "port_return": port_ann_return,
        "bm_return": bm_ann_return,
        "alpha": alpha,
        "tracking_error": tracking_error,
        "ir": ir,
        "correlation": correlation,
        "beta": beta,
    }


def optimize_portfolio(
        objective,
        mu,
        cov,
        current_w,
        bounds,
        group_map,
        group_limits,
        dur_array,
        yld_array,
        dur_limits,
        yld_limits,
        turnover_limit,
        returns_matrix,
        mar,
        rf,
        bm_returns_monthly=None,
        bm_dur=None,
        bm_yld=None,
        dur_diff_limits=None,
        yld_diff_limits=None,
):
    n_assets = len(current_w)
    eq_weights = np.ones(n_assets) / n_assets

    constraints = [{"type": "eq", "fun": lambda w: np.sum(w) - 1}]
    
    # 듀레이션 절대값 제약조건 (BM이 없는 경우에만 사용)
    if dur_limits is not None:
        constraints.append({"type": "ineq", "fun": lambda w: np.dot(w, dur_array) - dur_limits[0]})
        constraints.append({"type": "ineq", "fun": lambda w: dur_limits[1] - np.dot(w, dur_array)})
    
    # YLD 절대값 제약조건은 사용하지 않음
    # if yld_limits is not None:
    #     constraints.append({"type": "ineq", "fun": lambda w: np.dot(w, yld_array) - yld_limits[0]})
    #     constraints.append({"type": "ineq", "fun": lambda w: yld_limits[1] - np.dot(w, yld_array)})

    # BM 대비 차이 제약조건
    if bm_dur is not None and dur_diff_limits is not None:
        constraints.append({"type": "ineq", "fun": lambda w: np.dot(w, dur_array) - bm_dur - dur_diff_limits[0]})
        constraints.append({"type": "ineq", "fun": lambda w: dur_diff_limits[1] - (np.dot(w, dur_array) - bm_dur)})

    if bm_yld is not None and yld_diff_limits is not None:
        constraints.append({"type": "ineq", "fun": lambda w: np.dot(w, yld_array) - bm_yld - yld_diff_limits[0]})
        constraints.append({"type": "ineq", "fun": lambda w: yld_diff_limits[1] - (np.dot(w, yld_array) - bm_yld)})

    for group, idx in group_map.items():
        g_min, g_max = group_limits.get(group, (0, 1))
        constraints.append(
            {
                "type": "ineq",
                "fun": lambda w, indices=idx, g=g_min: np.sum(w[indices]) - g,
            }
        )
        constraints.append(
            {
                "type": "ineq",
                "fun": lambda w, indices=idx, g=g_max: g - np.sum(w[indices]),
            }
        )

    if turnover_limit is not None and turnover_limit > 0:
        constraints.append(
            {
                "type": "ineq",
                "fun": lambda w: turnover_limit - np.sum(np.abs(w - current_w)),
            }
        )

    # returns_matrix가 DataFrame인지 확인
    if isinstance(returns_matrix, pd.DataFrame):
        returns_matrix_values = returns_matrix.values
        returns_matrix_index = returns_matrix.index
    else:
        returns_matrix_values = returns_matrix
        returns_matrix_index = None

    def obj_fn(w):
        if objective == "Max Sharpe":
            vol = np.sqrt(np.dot(w, cov @ w))
            ret = np.dot(w, mu)
            if vol == 0:
                return 1e6
            return -((ret - rf) / vol)
        if objective == "Max Sortino":
            return -sortino_ratio(w, returns_matrix_values, mar, rf)
        if objective == "Max Return":
            return -np.dot(w, mu)
        if objective == "Min Risk":
            return np.sqrt(np.dot(w, cov @ w))
        if objective == "Risk Parity":
            return risk_parity_objective(w, cov)
        if objective == "Equal Weight":
            return np.sum((w - eq_weights) ** 2)
        if objective == "Max IR":
            # Information Ratio = Alpha / Tracking Error
            if bm_returns_monthly is None or len(bm_returns_monthly) == 0:
                return 1e6
            if returns_matrix_index is None:
                return 1e6
            port_monthly = pd.Series(returns_matrix_values @ w, index=returns_matrix_index)
            common_dates = returns_matrix_index.intersection(bm_returns_monthly.index)
            if len(common_dates) == 0:
                return 1e6
            port_common = port_monthly.loc[common_dates]
            bm_common = bm_returns_monthly.loc[common_dates]
            active_returns = port_common - bm_common
            tracking_error = np.std(active_returns) * np.sqrt(12)
            port_ann_return = (1 + port_common.mean()) ** 12 - 1
            bm_ann_return = (1 + bm_common.mean()) ** 12 - 1
            alpha = port_ann_return - bm_ann_return
            if tracking_error == 0:
                return 1e6
            ir = alpha / tracking_error
            return -ir  # 최대화를 위해 음수
        return 0.0

    result = minimize(
        obj_fn,
        current_w,
        method="SLSQP",
        bounds=bounds,
        constraints=constraints,
        options={"maxiter": 2000, "ftol": 1e-9, "disp": False},
    )

    if not result.success:
        st.warning(f"{objective} 최적화가 수렴하지 않아 현재 비중을 사용합니다.")
        optimized = current_w.copy()
    else:
        optimized = result.x

    optimized = optimized / np.sum(optimized)
    # returns_matrix가 DataFrame인지 확인
    returns_matrix_for_stats = returns_matrix_values if isinstance(returns_matrix, pd.DataFrame) else returns_matrix
    stats = portfolio_stats(optimized, mu, cov, dur_array, yld_array, returns_matrix_for_stats, mar, rf)
    turnover = np.sum(np.abs(optimized - current_w))

    return optimized, stats, turnover


def format_percentage(value):
    return f"{value * 100:.2f}%"


def format_ratio(value):
    if np.isnan(value):
        return "-"
    return f"{value:.2f}"


def main():
    st.title("Asset Allocation Optimizer")
    st.caption("Current / Asset_MinMax / Gr_MinMax / 기준가 / BM 시트를 포함한 엑셀을 업로드하세요.")

    # 예제 파일 다운로드 링크
    import os
    example_file_path = "images/Streamlit_Upload.xlsx"
    if os.path.exists(example_file_path):
        with open(example_file_path, "rb") as f:
            st.download_button(
                label="📥 예제 엑셀 파일 다운로드",
                data=f.read(),
                file_name="example.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    uploaded = st.file_uploader("엑셀 파일 선택", type=["xlsx", "xls"])
    if uploaded is None:
        st.stop()

    try:
        price_df, current_df_raw, asset_minmax_df, group_df, bm_df = load_excel(uploaded)
    except Exception as exc:
        st.error(f"엑셀 파일을 읽는 중 오류가 발생했습니다: {exc}")
        st.stop()

    # 사이드바 최상단에 이미지 표시
    try:
        st.sidebar.image("images/miraeasset.png", use_container_width=True)
    except:
        st.sidebar.warning("이미지를 불러올 수 없습니다.")

    # BM 사용 여부 선택
    use_bm = st.sidebar.radio(
        "BM 사용 여부",
        ("사용 안함", "사용"),
        index=0,
    )

    has_bm = (use_bm == "사용")

    # GD/GD+ 선택
    gd_option = st.sidebar.radio(
        "GD or GD+",
        ("GD", "GD+"),
        index=0,
    )

    # BM 사용 시 BM 시트 검증
    if has_bm:
        if bm_df is None:
            st.error("DATE, TR, DUR, YLD 칼럼을 가진 BM 시트가 입력파일에 존재해야 합니다.")
            st.stop()

        # BM 시트 칼럼 검증
        required_cols = ["DATE", "TR", "DUR", "YLD"]
        missing_cols = [col for col in required_cols if col not in bm_df.columns]
        if missing_cols:
            st.error(f"BM 시트에 다음 칼럼이 없습니다: {', '.join(missing_cols)}")
            st.stop()

    # normalize_percentage 적용 전 원본 값 확인 (5.0 또는 500% 체크용)
    raw_expected_r = current_df_raw.set_index("CODE")[
        "EXPECTED_R"] if "EXPECTED_R" in current_df_raw.columns else pd.Series()

    current_df = prepare_current_df(current_df_raw)
    group_df_prepared = prepare_group_df(group_df, gd_option)
    
    # 데이터 병합 (GD 옵션에 따라)
    current_df = merge_asset_data(current_df, asset_minmax_df, group_df_prepared, gd_option)

    codes = current_df["CODE"].astype(str).tolist()
    try:
        validate_codes(price_df, codes)
    except ValueError as exc:
        st.error(str(exc))
        st.stop()

    reference_months = st.sidebar.number_input("참조기간 (개월)", min_value=12, max_value=120, value=36, step=3)
    turnover_limit = st.sidebar.number_input("회전율 제약 (%)", min_value=0.0, max_value=200.0, value=100.0, step=5.0) / 100
    risk_free_rate = st.sidebar.number_input("무위험수익률 (%)", value=0.0, step=0.1) / 100
    mar = st.sidebar.number_input("Sortino 최소수익률 (%)", value=0.0, step=0.25) / 100
    mc_sims = st.sidebar.number_input("몬테칼로 시뮬레이션 횟수", min_value=100, max_value=5000, value=1000, step=100)
    expected_method = st.sidebar.radio(
        "기대수익률 산출 방식",
        ("excel", "historical", "monte_carlo"),
        format_func=lambda x: {"excel": "엑셀 입력", "historical": "과거수익률", "monte_carlo": "몬테칼로"}[x],
    )

    # 엑셀입력 방식을 선택한 경우 대체 방식 선택
    replacement_method = None
    if expected_method == "excel":
        replacement_method = st.sidebar.radio(
            "기대수익률 없을 경우 대체 방법",
            ("historical", "monte_carlo"),
            index=0,  # default: 과거수익률
            format_func=lambda x: {"historical": "과거수익률", "monte_carlo": "몬테칼로"}[x],
        )

    st.subheader("현재 자산 정보")
    # 표시용 컬럼 선택 (CODE 제외)
    display_cols = ["NAME", "CURRENT", "MIN", "MAX"]
    if "DUR" in current_df.columns:
        display_cols.append("DUR")
    if "YLD" in current_df.columns:
        display_cols.append("YLD")
    if "EXPECTED_R" in current_df.columns:
        display_cols.append("EXPECTED_R")
    if "GROUP" in current_df.columns:
        display_cols.append("GROUP")
    if "Gr_Min" in current_df.columns:
        display_cols.append("Gr_Min")
    if "Gr_Max" in current_df.columns:
        display_cols.append("Gr_Max")
    
    current_df_display = current_df[display_cols].copy()
    current_df_display = format_display_df(current_df_display, raw_expected_r)
    
    # 스타일링: NAME(Sleeve)를 제외한 나머지 칼럼은 오른쪽 정렬 + 1칸 들여쓰기
    def style_dataframe(df):
        styles = []
        for col in df.columns:
            if col != "Sleeve":
                styles.append({
                    'selector': f'th.col_{col}',
                    'props': [('text-align', 'right'), ('padding-right', '1em')]
                })
                styles.append({
                    'selector': f'td.col_{col}',
                    'props': [('text-align', 'right'), ('padding-right', '1em')]
                })
        return styles
    
    # Streamlit dataframe에 스타일 적용
    st.markdown("""
    <style>
        /* 모든 테이블 셀에 대해 Sleeve 칼럼을 제외하고 오른쪽 정렬 */
        div[data-testid="stDataFrame"] table thead tr th:not(:first-child),
        div[data-testid="stDataFrame"] table tbody tr td:not(:first-child),
        div[data-testid="stDataFrame"] table th:not(:first-child),
        div[data-testid="stDataFrame"] table td:not(:first-child) {
            text-align: right !important;
            padding-right: 1em !important;
        }
        /* 더 구체적인 선택자로 확실하게 적용 */
        div[data-testid="stDataFrame"] table th[data-testid*="columnHeader"]:not(:first-child),
        div[data-testid="stDataFrame"] table td:not(:first-child) {
            text-align: right !important;
            padding-right: 1em !important;
        }
        /* 모든 가능한 테이블 구조에 대해 적용 */
        div[data-testid="stDataFrame"] table tr th:nth-child(n+2),
        div[data-testid="stDataFrame"] table tr td:nth-child(n+2) {
            text-align: right !important;
            padding-right: 1em !important;
        }
        /* 테이블 스크롤 제거 - 모든 행이 한번에 보이도록 */
        div[data-testid="stDataFrame"] > div {
            max-height: none !important;
            overflow: visible !important;
        }
        div[data-testid="stDataFrame"] > div > div {
            max-height: none !important;
            overflow: visible !important;
        }
        div[data-testid="stDataFrame"] {
            max-height: none !important;
            overflow: visible !important;
        }
    </style>
    """, unsafe_allow_html=True)
    
    st.dataframe(current_df_display, use_container_width=True, hide_index=True)

    try:
        returns_df = compute_monthly_returns(price_df, codes, reference_months)
    except ValueError as exc:
        st.error(str(exc))
        st.stop()

    ann_returns = annualize_series(returns_df.mean())
    cov_monthly = returns_df.cov()
    cov_annual = cov_monthly * 12
    ann_vols = annualize_vol(np.sqrt(np.diag(cov_monthly)))
    ann_vols = pd.Series(ann_vols, index=returns_df.columns)

    # BM 수익률 및 DUR, YLD 계산
    bm_returns_monthly = None
    bm_dur = None
    bm_yld = None
    if has_bm:
        bm_returns_monthly = compute_bm_returns(bm_df, reference_months)
        if bm_returns_monthly is None:
            st.error("BM 시트에서 TR 데이터를 읽을 수 없습니다.")
            st.stop()

        # BM DUR, YLD 계산 (최신 값 사용)
        bm_df_sorted = bm_df.copy()
        bm_df_sorted["DATE"] = pd.to_datetime(bm_df_sorted["DATE"], errors="coerce")
        bm_df_sorted = bm_df_sorted.sort_values("DATE")
        bm_dur = bm_df_sorted["DUR"].iloc[-1] if len(bm_df_sorted) > 0 and not pd.isna(
            bm_df_sorted["DUR"].iloc[-1]) else None
        bm_yld = bm_df_sorted["YLD"].iloc[-1] if len(bm_df_sorted) > 0 and not pd.isna(
            bm_df_sorted["YLD"].iloc[-1]) else None

        if bm_dur is None or bm_yld is None:
            st.error("BM 시트에서 DUR 또는 YLD 값을 읽을 수 없습니다.")
            st.stop()

    # 엑셀에서 기대수익률 값 확인
    # 엑셀에서 % 형식으로 500% 입력 시 pandas가 5.0으로 읽을 수 있으므로, 원본 값과 정규화된 값 모두 체크
    excel_expected_r = current_df.set_index("CODE")["EXPECTED_R"] if "EXPECTED_R" in current_df.columns else pd.Series(dtype=float, index=current_df.set_index("CODE").index)
    
    # 원본 값에서 5.0 또는 500을 체크, 또는 정규화된 값에서 5.0을 체크
    # (엑셀에서 일반 형식으로 500 입력 → normalize_percentage에서 500/100 = 5.0)
    # (엑셀에서 % 형식으로 500% 입력 → pandas가 5.0으로 읽음 → normalize_percentage에서 5.0/100 = 0.05)
    raw_r_aligned = raw_expected_r.reindex(excel_expected_r.index) if len(raw_expected_r) > 0 else pd.Series(
        index=excel_expected_r.index)
    missing_r = excel_expected_r.isna() | np.isclose(excel_expected_r, 5.0, atol=0.1) | (raw_r_aligned.notna() & (
                np.isclose(raw_r_aligned, 5.0, atol=0.1) | np.isclose(raw_r_aligned, 500.0, atol=1.0)))

    # 모든 자산이 기대수익률이 없고 엑셀입력을 선택한 경우
    if expected_method == "excel" and missing_r.all():
        st.error("모든 자산의 기대수익률이 없습니다. 엑셀입력 방식을 사용할 수 없습니다. 과거수익률 또는 몬테칼로 방식을 선택해주세요.")
        st.stop()

    expected_returns = build_expected_returns(
        expected_method,
        current_df,
        ann_returns,
        returns_df.mean(),
        cov_monthly,
        mc_sims,
        raw_expected_r=raw_expected_r if len(raw_expected_r) > 0 else None,
        replacement_method=replacement_method,
    )
    expected_vols = build_expected_vols(
        expected_method,
        current_df,
        ann_vols,
        returns_df.mean(),
        cov_monthly,
        mc_sims,
        raw_expected_v=None,
        replacement_method=replacement_method,
    )

    # 메시지 처리
    if expected_method == "excel" and missing_r.any():
        # 일부 자산만 기대수익률이 없는 경우
        missing_codes = current_df.loc[missing_r.values, "NAME"].tolist()
        missing_names = ", ".join(missing_codes)
        method_name = "과거수익률" if replacement_method == "historical" else "몬테칼로"
        st.warning(
            f"{missing_names} 자산은 기대수익률이 입력되지 않았습니다. 그래서 사이드바에서 선택한 방식({method_name})으로 값을 대체합니다.")
    elif expected_method in ["historical", "monte_carlo"] and missing_r.all():
        # 모든 자산이 기대수익률이 없고 과거수익률/몬테칼로를 선택한 경우
        method_name = "과거수익률" if expected_method == "historical" else "몬테칼로"
        st.info(f"모든 자산의 기대수익률이 없어 선택한 방식({method_name})으로 값을 대체합니다.")

    # BM 정보 표시
    st.subheader("벤치마크 (BM) 정보")
    if not has_bm:
        st.info("⚠️ BM 사용 안함을 선택했습니다. BM이 없는 최적화를 수행합니다.")
    else:
        # BM 수익률 정보 표시
        bm_ann_return = (1 + bm_returns_monthly.mean()) ** 12 - 1
        bm_ann_vol = np.std(bm_returns_monthly) * np.sqrt(12)
        bm_info = pd.DataFrame({
            "항목": ["연율 기대수익률", "연율 변동성", "DUR", "YLD", "데이터 기간"],
            "값": [
                f"{bm_ann_return * 100:.2f}%",
                f"{bm_ann_vol * 100:.2f}%",
                f"{bm_dur:.2f}",
                f"{bm_yld * 100:.2f}%",
                f"{len(bm_returns_monthly)}개월"
            ]
        })
        st.dataframe(bm_info, use_container_width=True, hide_index=True)

    current_weights = current_df["CURRENT"].fillna(0).values
    if current_weights.sum() == 0:
        st.error("Current 시트의 CURRENT 컬럼 합계가 0입니다.")
        st.stop()
    current_weights = current_weights / current_weights.sum()

    dur_array = current_df["DUR"].fillna(0).values
    yld_array = current_df["YLD"].fillna(0).values

    current_duration = np.dot(current_weights, dur_array)
    current_yield = np.dot(current_weights, yld_array)
    current_stats = portfolio_stats(
        current_weights,
        expected_returns.values,
        cov_annual.values,
        dur_array,
        yld_array,
        returns_df.values,
        mar,
        risk_free_rate,
    )

    st.sidebar.markdown("---")
    st.sidebar.markdown("### 현재 포트폴리오 지표")
    st.sidebar.markdown(f"**현재 듀레이션:** {current_duration:.2f}")
    st.sidebar.markdown(f"**현재 YLD:** {current_yield * 100:.2f}%")
    if has_bm:
        st.sidebar.markdown(f"**BM 듀레이션:** {bm_dur:.2f}")
        st.sidebar.markdown(f"**BM YLD:** {bm_yld * 100:.2f}%")
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 제약조건 설정")

    if has_bm:
        # BM 대비 차이 제약조건만 사용
        st.sidebar.markdown("#### BM 대비 차이 제약")
        dur_diff_min = st.sidebar.number_input("BM 대비 DUR 차이 최소", value=-1.0, step=0.1)
        dur_diff_max = st.sidebar.number_input("BM 대비 DUR 차이 최대", value=1.0, step=0.1)
        yld_diff_min = st.sidebar.number_input("BM 대비 YLD 차이 최소 (%)", value=-0.5, step=0.1) / 100
        yld_diff_max = st.sidebar.number_input("BM 대비 YLD 차이 최대 (%)", value=0.5, step=0.1) / 100
        # BM이 있는 경우 절대 제약은 사용하지 않음
        dur_min = None
        dur_max = None
    else:
        dur_diff_min = None
        dur_diff_max = None
        yld_diff_min = None
        yld_diff_max = None
        dur_min_pct = st.sidebar.number_input("듀레이션 최소 (%)", value=-20.0, step=1.0, help="현재 듀레이션 대비 하한 비율")
        dur_max_pct = st.sidebar.number_input("듀레이션 최대 (%)", value=20.0, step=1.0, help="현재 듀레이션 대비 상한 비율")
        dur_min = current_duration * (1 + dur_min_pct / 100)
        dur_max = current_duration * (1 + dur_max_pct / 100)
        st.sidebar.caption(f"듀레이션 범위: {dur_min:.2f} ~ {dur_max:.2f}")

    bounds = list(
        zip(
            current_df["MIN"].fillna(0).values,
            current_df["MAX"].fillna(1).values,
        )
    )

    # 병합된 데이터에서 GROUP별 Gr_Min, Gr_Max를 사용하여 group_limits 생성
    group_limits = {}
    if "GROUP" in current_df.columns and "Gr_Min" in current_df.columns and "Gr_Max" in current_df.columns:
        # 각 GROUP의 첫 행에서 Gr_Min, Gr_Max를 가져옴
        for group in current_df["GROUP"].unique():
            if pd.isna(group):
                continue
            group_rows = current_df[current_df["GROUP"] == group]
            if len(group_rows) > 0:
                gr_min = group_rows.iloc[0]["Gr_Min"]
                gr_max = group_rows.iloc[0]["Gr_Max"]
                # NaN이면 기본값 사용
                gr_min = gr_min if not pd.isna(gr_min) else 0
                gr_max = gr_max if not pd.isna(gr_max) else 1
                group_limits[group] = (gr_min, gr_max)
    else:
        # Gr_Min, Gr_Max가 없으면 기본값 사용
        for group in current_df["GROUP"].unique() if "GROUP" in current_df.columns else []:
            if pd.isna(group):
                continue
            group_limits[group] = (0, 1)
    group_map = {}
    for idx, row in current_df.reset_index().iterrows():
        group = row.get("GROUP")
        if pd.isna(group):
            continue
        group_map.setdefault(group, []).append(idx)

    objectives = [
        "Max Sharpe",
        "Max Sortino",
        "Max Return",
        "Min Risk",
        "Risk Parity",
        "Equal Weight",
    ]

    # BM이 있으면 MaxIR 추가
    if has_bm:
        objectives.append("Max IR")

    # session_state 초기화
    if "optimization_results" not in st.session_state:
        st.session_state.optimization_results = None
    if "optimization_objectives" not in st.session_state:
        st.session_state.optimization_objectives = None
    if "optimization_current_df" not in st.session_state:
        st.session_state.optimization_current_df = None
    if "optimization_current_weights" not in st.session_state:
        st.session_state.optimization_current_weights = None
    if "optimization_current_stats" not in st.session_state:
        st.session_state.optimization_current_stats = None

    if st.button("최적화 실행"):
        progress_placeholder = st.empty()
        progress_placeholder.info("최적화 진행 중...")

        results = {}
        for obj in objectives:
            opt_w, stats, turnover = optimize_portfolio(
                obj,
                expected_returns.values,
                cov_annual.values,
                current_weights,
                bounds,
                group_map,
                group_limits,
                dur_array,
                yld_array,
                (dur_min, dur_max) if dur_min is not None else None,  # BM이 있는 경우 None
                None,  # YLD 절대값 제약조건 사용 안함
                turnover_limit,
                returns_df,
                mar,
                risk_free_rate,
                bm_returns_monthly=bm_returns_monthly,
                bm_dur=bm_dur if has_bm else None,
                bm_yld=bm_yld if has_bm else None,
                dur_diff_limits=(dur_diff_min, dur_diff_max) if has_bm else None,
                yld_diff_limits=(yld_diff_min, yld_diff_max) if has_bm else None,
            )

            # 포트폴리오 vs BM 비교 지표 계산
            bm_stats = None
            if has_bm:
                bm_stats = portfolio_vs_bm_stats(opt_w, expected_returns.values, cov_annual.values, returns_df,
                                                 bm_returns_monthly)

            results[obj] = {
                "weights": opt_w,
                "stats": stats,
                "turnover": turnover,
                "bm_stats": bm_stats,
            }

        # 최적화 완료 후 진행 메시지 제거
        progress_placeholder.empty()

        # 결과를 session_state에 저장
        st.session_state.optimization_results = results
        st.session_state.optimization_objectives = objectives
        st.session_state.optimization_current_df = current_df
        st.session_state.optimization_current_weights = current_weights
        st.session_state.optimization_current_stats = current_stats

    # 저장된 결과가 있으면 표시
    if st.session_state.optimization_results is not None:
        results = st.session_state.optimization_results
        objectives = st.session_state.optimization_objectives
        current_df = st.session_state.optimization_current_df
        current_weights = st.session_state.optimization_current_weights
        current_stats = st.session_state.optimization_current_stats

        # 엑셀 내보내기 함수
        def create_excel_file():
            import io
            try:
                from openpyxl import Workbook
                from openpyxl.utils.dataframe import dataframe_to_rows
                
                excel_buffer = io.BytesIO()
                wb = Workbook()
                wb.remove(wb.active)  # 기본 시트 제거
                
                # 기본 테이블 구조 (모든 자산 포함)
                base_table = current_df[["NAME"]].copy()
                
                # 각 최적화 목표별로 시트 생성 (모든 행 포함)
                for obj in objectives:
                    opt = results[obj]["weights"]
                    table_export = base_table.copy()
                    table_export["Current Weight"] = [format_percentage(w) for w in current_weights]
                    table_export["Optimized Weight"] = [format_percentage(w) for w in opt]
                    table_export["Diff"] = [format_percentage(w) for w in (opt - current_weights)]
                    
                    # 시트 이름 생성 (최대 31자)
                    sheet_name = obj[:31] if len(obj) > 31 else obj
                    ws = wb.create_sheet(title=sheet_name)
                    
                    # 데이터프레임을 시트에 추가
                    for r in dataframe_to_rows(table_export, index=False, header=True):
                        ws.append(r)
                
                # 방법론별 optimized weight만 있는 시트 생성 (Current Weight 포함)
                summary_table = base_table.copy()
                summary_table["Current Weight"] = [format_percentage(w) for w in current_weights]
                for obj in objectives:
                    opt = results[obj]["weights"]
                    # 시트 이름을 컬럼명으로 사용 (최대 31자)
                    col_name = obj[:31] if len(obj) > 31 else obj
                    summary_table[col_name] = [format_percentage(w) for w in opt]
                
                ws_summary = wb.create_sheet(title="Summary", index=0)  # 첫 번째 시트로
                for r in dataframe_to_rows(summary_table, index=False, header=True):
                    ws_summary.append(r)
                
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                return excel_buffer.getvalue()
            except ImportError:
                return None
        
        # 엑셀 파일 생성 및 다운로드 버튼 표시
        excel_data = create_excel_file()
        if excel_data:
            st.download_button(
                label="📥 최적화 결과 엑셀 다운로드",
                data=excel_data,
                file_name="optimization_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("엑셀 내보내기를 위해 openpyxl 패키지가 필요합니다. pip install openpyxl로 설치해주세요.")

        tabs = st.tabs(objectives)
        for tab, obj in zip(tabs, objectives):
            with tab:
                opt = results[obj]["weights"]
                stats = results[obj]["stats"]
                turnover = results[obj]["turnover"]

                table = current_df[["NAME"]].copy()
                table = table.rename(columns={"NAME": "Sleeve"})
                table["Current Weight"] = [format_percentage(w) for w in current_weights]
                table["Optimized Weight"] = [format_percentage(w) for w in opt]
                table["Diff"] = [format_percentage(w) for w in opt - current_weights]
                st.dataframe(table, use_container_width=True, hide_index=True)

                st.markdown(
                    f"""
                    - 기대수익률: {format_percentage(stats['return'])}
                    - 변동성: {format_percentage(stats['vol'])}
                    - 샤프지수: {format_ratio(stats['sharpe'])}
                    - 소르티노지수: {format_ratio(stats['sortino'])}
                    - 듀레이션: {current_stats['duration']:.2f} -> {stats['duration']:.2f}
                    - YLD: {format_percentage(current_stats['yield'])} -> {format_percentage(stats['yield'])}
                    - 회전율: {format_percentage(turnover)}
                    """
                )

                # 포트폴리오 vs BM 비교 지표 표시
                if has_bm and results[obj]["bm_stats"] is not None:
                    bm_stats = results[obj]["bm_stats"]
                    st.markdown("---")
                    st.markdown("### 포트폴리오 vs 벤치마크 비교")
                    st.markdown(
                        f"""
                        - 포트폴리오 수익률: {format_percentage(bm_stats['port_return'])}
                        - 벤치마크 수익률: {format_percentage(bm_stats['bm_return'])}
                        - Alpha (초과수익률): {format_percentage(bm_stats['alpha'])}
                        - Tracking Error (TE): {format_percentage(bm_stats['tracking_error'])}
                        - Information Ratio (IR): {format_ratio(bm_stats['ir'])}
                        - 상관관계: {format_ratio(bm_stats['correlation'])}
                        - Beta: {format_ratio(bm_stats['beta'])}
                        """
                    )


if __name__ == "__main__":
    main()



